"""
도시가스 가스공사 도매요금 정산 계산기
- 깃허브 레포에서 판매량정산서(xlsx) 자동 로드
- 월별 수급량(GJ) 입력 시 용도별 공급량·구성비 산출
"""

import os
import re
import requests
import openpyxl
import pandas as pd
import streamlit as st
from io import BytesIO
from typing import Optional

# ──────────────────────────────────────────
# 설정
# ──────────────────────────────────────────
GITHUB_USER = "Han11112222"          # ← 본인 깃허브 유저명
GITHUB_REPO = "gas-settlement"       # ← 본인 레포명 (README에서 확인)
GITHUB_BRANCH = "main"

# 용도 행 위치 (가스공사용(MJ) 시트 기준)
# row_index: 엑셀 행번호(1-based), label: 표시명, is_subtotal: 소계 여부
YONGDO_ROWS = [
    {"row": 9,  "label": "주택용(소계)",   "is_subtotal": True},
    {"row": 12, "label": "일반용",         "is_subtotal": True},
    {"row": 13, "label": "냉난방공조용",   "is_subtotal": False},
    {"row": 14, "label": "업무난방용",     "is_subtotal": False},
    {"row": 15, "label": "산업용",         "is_subtotal": False},
    {"row": 16, "label": "수송용",         "is_subtotal": False},
    {"row": 17, "label": "열병합용",       "is_subtotal": False},
    {"row": 18, "label": "연료전지용",     "is_subtotal": False},
    {"row": 19, "label": "열전용설비용",   "is_subtotal": False},
    {"row": 20, "label": "주한미군",       "is_subtotal": False},
    {"row": 21, "label": "합계",           "is_subtotal": True},
]

# L열(판매량계) = 인덱스 11 (0-based), M열(구성비) = 인덱스 12
COL_SALES_TOTAL = 11   # L
COL_RATIO       = 12   # M


# ──────────────────────────────────────────
# GitHub 파일 목록 & 다운로드
# ──────────────────────────────────────────
@st.cache_data(ttl=300)
def get_github_file_list() -> list[str]:
    """레포 루트의 파일 목록을 가져옴"""
    url = f"https://api.github.com/repos/{GITHUB_USER}/{GITHUB_REPO}/git/trees/{GITHUB_BRANCH}?recursive=1"
    headers = {}
    token = os.environ.get("GITHUB_TOKEN", "")
    if token:
        headers["Authorization"] = f"token {token}"
    r = requests.get(url, headers=headers, timeout=10)
    if r.status_code != 200:
        st.error(f"GitHub API 오류: {r.status_code} – {r.text[:200]}")
        return []
    tree = r.json().get("tree", [])
    return [item["path"] for item in tree if item["path"].endswith(".xlsx")]


@st.cache_data(ttl=300)
def download_xlsx(filename: str) -> Optional[bytes]:
    """raw 파일 다운로드"""
    url = f"https://raw.githubusercontent.com/{GITHUB_USER}/{GITHUB_REPO}/{GITHUB_BRANCH}/{requests.utils.quote(filename)}"
    headers = {}
    token = os.environ.get("GITHUB_TOKEN", "")
    if token:
        headers["Authorization"] = f"token {token}"
    r = requests.get(url, headers=headers, timeout=15)
    if r.status_code == 200:
        return r.content
    return None


def parse_yearmonth(filename: str) -> Optional[tuple[int, int]]:
    """
    파일명에서 (연도, 월) 추출
    예: 판매량정산서2025년08월(확정).xlsx → (2025, 8)
        판매량정산서2026년5월(확정).xlsx  → (2026, 5)
    """
    m = re.search(r"(\d{4})년(\d{1,2})월", filename)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


# ──────────────────────────────────────────
# 엑셀 파싱
# ──────────────────────────────────────────
def read_gasco_sheet(xlsx_bytes: bytes) -> dict:
    """
    '가스공사용(MJ)' 시트에서 용도별 판매량계(MJ)·구성비(%)를 읽어 반환
    반환: {label: {"sales_mj": float, "ratio": float}}
    """
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if "가스공사용(MJ)" not in wb.sheetnames:
        return {}
    ws = wb["가스공사용(MJ)"]
    all_rows = list(ws.iter_rows(values_only=True))

    result = {}
    for yongdo in YONGDO_ROWS:
        row_idx = yongdo["row"] - 1   # 0-based
        if row_idx >= len(all_rows):
            continue
        row = all_rows[row_idx]
        sales = row[COL_SALES_TOTAL]
        ratio = row[COL_RATIO]
        if isinstance(sales, (int, float)) and sales:
            result[yongdo["label"]] = {
                "sales_mj": float(sales),
                "ratio":    float(ratio) if isinstance(ratio, (int, float)) else None,
                "is_subtotal": yongdo["is_subtotal"],
            }
    return result


# ──────────────────────────────────────────
# 정산 계산
# ──────────────────────────────────────────
def calc_settlement(
    supply_gj: float,
    prev_data: dict,
    curr_data: dict,
    method: str = "당월단독",
) -> pd.DataFrame:
    """
    용도별 배분량(GJ) 계산
    method: "당월단독" | "전월평균"
    """
    labels = [r["label"] for r in YONGDO_ROWS]
    rows = []

    total_prev = prev_data.get("합계", {}).get("sales_mj", None)
    total_curr = curr_data.get("합계", {}).get("sales_mj", None)

    for label in labels:
        prev = prev_data.get(label, {})
        curr = curr_data.get(label, {})

        prev_mj = prev.get("sales_mj")
        curr_mj = curr.get("sales_mj")

        # 구성비 계산
        if method == "전월평균" and prev_mj and curr_mj and total_prev and total_curr:
            r_prev = prev_mj / total_prev * 100
            r_curr = curr_mj / total_curr * 100
            ratio = (r_prev + r_curr) / 2
        elif curr_mj and total_curr:
            ratio = curr_mj / total_curr * 100
        elif prev_mj and total_prev:
            ratio = prev_mj / total_prev * 100
        else:
            ratio = None

        alloc_gj = supply_gj * ratio / 100 if ratio is not None else None

        is_sub = (prev or curr).get("is_subtotal", False)
        rows.append({
            "용도":           label,
            "구성비(%)":      round(ratio, 4) if ratio else None,
            "배분공급량(GJ)": round(alloc_gj, 3) if alloc_gj else None,
            "비고":           "소계" if is_sub else "",
        })

    return pd.DataFrame(rows)


# ──────────────────────────────────────────
# Streamlit UI
# ──────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="도시가스 도매요금 정산 계산기",
        page_icon="🔥",
        layout="wide",
    )

    # ── 헤더
    st.markdown(
        """
        <style>
        .main-title {
            font-size: 1.6rem; font-weight: 700;
            color: #1a3a5c; border-bottom: 2px solid #e65c00;
            padding-bottom: 0.3rem; margin-bottom: 0.2rem;
        }
        .sub-title { font-size: 0.9rem; color: #666; margin-bottom: 1.2rem; }
        .highlight { background: #fff7f0; border-left: 3px solid #e65c00;
                     padding: 0.5rem 0.8rem; border-radius: 4px; font-size: 0.85rem; }
        </style>
        <div class="main-title">🔥 도시가스 도매요금 정산 계산기</div>
        <div class="sub-title">판매량정산서 기반 용도별 구성비 & 공급량 자동 산출</div>
        """,
        unsafe_allow_html=True,
    )

    # ── GitHub 파일 목록 로드
    with st.spinner("GitHub에서 판매량정산서 목록 로딩 중…"):
        xlsx_files = get_github_file_list()

    if not xlsx_files:
        st.warning("GitHub 레포에서 xlsx 파일을 찾지 못했습니다. GITHUB_USER / GITHUB_REPO 설정을 확인하세요.")
        return

    # 파일 → (연도, 월) 매핑
    file_map: dict[tuple[int, int], str] = {}
    for f in xlsx_files:
        ym = parse_yearmonth(f)
        if ym:
            file_map[ym] = f

    if not file_map:
        st.error("연도·월 패턴을 가진 xlsx 파일이 없습니다.")
        return

    sorted_keys = sorted(file_map.keys())
    month_labels = [f"{y}년 {m:02d}월" for y, m in sorted_keys]

    # ── 입력 패널
    st.markdown("---")
    col1, col2, col3 = st.columns([2, 2, 2])

    with col1:
        st.markdown("**① 정산 대상 월 선택**")
        selected_label = st.selectbox("정산월", month_labels, index=len(month_labels) - 1)
        selected_ym = sorted_keys[month_labels.index(selected_label)]

    with col2:
        st.markdown("**② 수급량 입력 (GJ)**")
        supply_gj = st.number_input(
            "월 수급량 (GJ)",
            min_value=0.0,
            value=1_774_554.307,
            step=1000.0,
            format="%.3f",
            help="가스공사 공급량 일일보고서의 '월' 수급량(GJ) 입력",
        )

    with col3:
        st.markdown("**③ 구성비 산출 방식**")
        method = st.radio(
            "방식",
            ["당월단독", "전월평균"],
            help="당월단독: 해당 월 판매량만 사용\n전월평균: (m-1)월 + m월 평균",
        )

    # ── 이전 월 자동 선택
    curr_idx = sorted_keys.index(selected_ym)
    prev_ym = sorted_keys[curr_idx - 1] if curr_idx > 0 else None

    st.markdown(
        f'<div class="highlight">📁 당월 파일: <b>{file_map[selected_ym]}</b>'
        + (f' &nbsp;|&nbsp; 전월 파일: <b>{file_map[prev_ym]}</b>' if prev_ym else "")
        + "</div>",
        unsafe_allow_html=True,
    )
    st.markdown("")

    # ── 계산 버튼
    if st.button("🔢 구성비 & 공급량 산출", type="primary", use_container_width=True):
        with st.spinner("엑셀 파일 다운로드 & 계산 중…"):

            # 당월 파일
            curr_bytes = download_xlsx(file_map[selected_ym])
            if not curr_bytes:
                st.error(f"당월 파일 다운로드 실패: {file_map[selected_ym]}")
                return
            curr_data = read_gasco_sheet(curr_bytes)

            # 전월 파일
            prev_data = {}
            if prev_ym and method == "전월평균":
                prev_bytes = download_xlsx(file_map[prev_ym])
                if prev_bytes:
                    prev_data = read_gasco_sheet(prev_bytes)

            # 계산
            df = calc_settlement(supply_gj, prev_data, curr_data, method)

        # ── 결과 표시
        st.markdown("### 📊 용도별 공급량 배분 결과")

        # 스타일링
        def style_row(row):
            if row["용도"] in ("합계", "주택용(소계)", "일반용"):
                return ["background-color: #f0f4fa; font-weight: bold"] * len(row)
            return [""] * len(row)

        main_df = df[df["용도"] != "합계"].copy()
        total_row = df[df["용도"] == "합계"].copy()

        styled = main_df.style.apply(style_row, axis=1).format({
            "구성비(%)":      lambda x: f"{x:.4f}%" if pd.notna(x) else "-",
            "배분공급량(GJ)": lambda x: f"{x:,.3f}" if pd.notna(x) else "-",
        })
        st.dataframe(styled, use_container_width=True, height=420)

        # 합계 강조
        if not total_row.empty:
            t = total_row.iloc[0]
            c1, c2, c3 = st.columns(3)
            c1.metric("수급량 (입력)", f"{supply_gj:,.3f} GJ")
            c2.metric("배분합계 (GJ)", f"{t['배분공급량(GJ)']:,.3f}" if pd.notna(t['배분공급량(GJ)']) else "-")
            c3.metric("구성비 합계", f"{t['구성비(%)']:.4f}%" if pd.notna(t['구성비(%)']) else "-")

        # ── 산출 근거 상세
        with st.expander("📋 산출 근거 상세 (판매량계 MJ)"):
            detail_rows = []
            total_curr = curr_data.get("합계", {}).get("sales_mj", 1)
            total_prev = prev_data.get("합계", {}).get("sales_mj", 1) if prev_data else None

            for yongdo in YONGDO_ROWS:
                label = yongdo["label"]
                curr = curr_data.get(label, {})
                prev = prev_data.get(label, {}) if prev_data else {}
                detail_rows.append({
                    "용도":              label,
                    f"당월 판매량계(MJ)": f"{curr.get('sales_mj', 0):,.0f}" if curr.get('sales_mj') else "-",
                    f"당월 구성비(%)":    f"{curr.get('sales_mj', 0)/total_curr*100:.4f}" if curr.get('sales_mj') else "-",
                    f"전월 판매량계(MJ)": f"{prev.get('sales_mj', 0):,.0f}" if prev.get('sales_mj') else "-",
                    f"전월 구성비(%)":    f"{prev.get('sales_mj', 0)/total_prev*100:.4f}" if prev.get('sales_mj') and total_prev else "-",
                })
            st.dataframe(pd.DataFrame(detail_rows), use_container_width=True)

        # ── CSV 다운로드
        csv = df.to_csv(index=False, encoding="utf-8-sig")
        st.download_button(
            label="⬇️ 결과 CSV 다운로드",
            data=csv.encode("utf-8-sig"),
            file_name=f"도매정산_{selected_label.replace(' ', '')}.csv",
            mime="text/csv",
        )

    # ── 사용 안내
    with st.expander("ℹ️ 사용 방법 & 로직 설명"):
        st.markdown(
            """
**로직 요약**

| 단계 | 내용 |
|------|------|
| ① | GitHub 레포에서 해당 월 판매량정산서 xlsx 자동 다운로드 |
| ② | `가스공사용(MJ)` 시트의 용도별 **판매량계(MJ)** 추출 |
| ③ | 판매량계 / 합계 × 100 = **용도별 구성비(%)** |
| ④ | 구성비 × 수급량(GJ) = **용도별 배분 공급량(GJ)** |

**구성비 산출 방식**
- **당월단독**: 해당 월 판매량만 사용 (일반적인 경우)
- **전월평균**: (m-1월 구성비 + m월 구성비) ÷ 2 (비교·검증용)

**판매량계 포함 항목**
- 실제 검침 판매량(MJ)
- 사회복지시설 할인 복원분
- 자가소모량 (열병합발전, 냉난방, 퍼지량 등)
- **제외**: 수송용 BIO (별도 바이오가스 정산)

**환경변수 설정 (private 레포의 경우)**
```
GITHUB_TOKEN=ghp_xxxxxxxxxxxx
```
            """
        )


if __name__ == "__main__":
    main()
